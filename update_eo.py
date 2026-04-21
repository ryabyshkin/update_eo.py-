#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EO Report Updater
Обновляет шаблон EO данными из файлов 1С, сохраняет EO_updated.xlsx.

Запуск:
    python3 update_eo.py

Все файлы должны лежать в одной папке со скриптом.
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

FOLDER = Path(__file__).parent

TEMPLATE      = FOLDER / 'EO_template.xlsx'
OUTPUT        = FOLDER / 'EO_updated.xlsx'
NEW_SALES_COL = 'Продажи 09.04.-15.04. шт'

SHEETS = ['SS26', 'FW25', 'БАЗА', 'Предыдущие дропы', 'Сопутка и прочее']

WAREHOUSES = {
    'В пути в контент отдел Москва Бережковская (В пути)',
    'В пути в магазин Красноярск (В пути)',
    'В пути в магазин Москва Авиапарк (В пути)',
    'В пути в магазин Новосибирск (В пути)',
    'В пути в магазин С-Петербург (В пути)',
    'В пути в магазин С. Вражек (В пути)',
    'В пути в магазин Томск (В пути)',
    'В пути на ГП Москва (Бережковская)',
    'Готовая продукция Москва (Бережковская)',
    'Готовая продукция Томск МАКСПРО',
    'Контент отдел Москва Бережковская',
    'Магазин Красноярск',
    'Магазин Москва Авиапарк',
    'Магазин Москва С. Вражек',
    'Магазин Новосибирск',
    'Магазин С-Петербург',
    'Магазин Томск',
}

FILL_A = PatternFill(fill_type='solid', fgColor='92D050')
FILL_B = PatternFill(fill_type='solid', fgColor='FFFF00')
FILL_C = PatternFill(fill_type='solid', fgColor='FF0000')

WEEKLY_RE = re.compile(r'^Продажи\s+\d{2}\.\d{2}\.-\d{2}\.\d{2}\.\s*шт$')

def build_key(nom, har):
    nom = str(nom).strip() if nom is not None else ''
    har = str(har).strip() if har is not None else ''
    har = har.split(';')[0].strip()
    return f'{nom} ({har})'

def to_num(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(str(v).replace(',', '.').replace('\xa0', '').replace('\u202f', '').replace(' ', ''))
    except (ValueError, TypeError):
        return 0.0

def norm(s):
    return str(s).strip().lower().replace('  ', ' ')

def find_col(ws, header_row, col_name):
    target = norm(col_name)
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=ci).value
        if v is not None and norm(v) == target:
            return ci
    return None

def get_header_row(ws):
    target = norm('Номенклатура+характеристика')
    for ri in range(1, 11):
        for ci in range(1, ws.max_column + 1):
            v = ws.cell(row=ri, column=ci).value
            if v is not None and norm(v) == target:
                return ri
    raise ValueError(f'«Номенклатура+характеристика» не найдена на листе «{ws.title}»')

def build_key_index(ws, header_row, key_col):
    idx = {}
    for ri in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=ri, column=key_col).value
        if v:
            idx[str(v).strip()] = ri
    return idx

def fill_column(ws, col_idx, data_map, key_index):
    filled = 0
    for key, ri in key_index.items():
        cell = ws.cell(row=ri, column=col_idx)
        if key in data_map:
            cell.value = data_map[key]
            filled += 1
        elif str(cell.value).strip() == 'нет данных':
            cell.value = None
    return filled

def read_vypusk():
    print('  vypusk_shi.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'vypusk_shi.xlsx', data_only=True).active
    m = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        nom = str(row[0]).strip() if row[0] else ''
        if not nom or nom == 'None': continue
        k = build_key(nom, row[2])
        m[k] = m.get(k, 0) + to_num(row[4])
    print(f'    → {len(m)} ключей')
    return m

def read_opt():
    print('  opt.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'opt.xlsx', data_only=True).active
    z, o = {}, {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        nom = str(row[0]).strip() if row[0] else ''
        if not nom or nom == 'None': continue
        k = build_key(nom, row[2])
        z[k] = z.get(k, 0) + to_num(row[7])
        o[k] = o.get(k, 0) + to_num(row[8])
    print(f'    → {len(z)} ключей')
    return z, o

def read_ostatok():
    print('  ostatok.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'ostatok.xlsx', data_only=True).active
    all_rows = list(ws.iter_rows(values_only=True))
    wh_cols = []
    for hi in range(min(4, len(all_rows))):
        for ci, val in enumerate(all_rows[hi]):
            if val and str(val).strip() in WAREHOUSES and ci not in wh_cols:
                wh_cols.append(ci)
    print(f'    складов найдено: {len(wh_cols)}/{len(WAREHOUSES)}')
    m = {}
    for row in all_rows[4:]:
        nom = str(row[3]).strip() if row[3] else ''
        if not nom or nom == 'None': continue
        k = build_key(nom, row[5])
        total = sum(to_num(row[ci]) for ci in wh_cols if ci < len(row))
        m[k] = m.get(k, 0) + total
    print(f'    → {len(m)} ключей')
    return m

def read_prodazhi():
    print('  prodazhi_nedelya.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'prodazhi_nedelya.xlsx', data_only=True).active
    all_rows = list(ws.iter_rows(values_only=True))
    shts, rub = {}, {}
    for row in all_rows[20:]:
        har = str(row[1]).strip() if row[1] else ''
        if not har or har == 'None': continue
        nom = str(row[0]).strip() if row[0] else ''
        if not nom or nom == 'None': continue
        k = build_key(nom, har)
        v21 = to_num(row[21]) if len(row) > 21 else 0
        v28 = to_num(row[28]) if len(row) > 28 else 0
        v22 = to_num(row[22]) if len(row) > 22 else 0
        v29 = to_num(row[29]) if len(row) > 29 else 0
        shts[k] = shts.get(k, 0) + (v21 - v28)
        rub[k]  = rub.get(k,  0) + (v22 - v29)
    print(f'    → {len(shts)} ключей')
    return shts, rub

def read_rezerv_lamoda():
    print('  rezerv_lamoda.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'rezerv_lamoda.xlsx', data_only=True).active
    all_rows = list(ws.iter_rows(values_only=True))
    m = {}
    for row in all_rows[1:]:
        nom = str(row[0]).strip() if row[0] else ''
        har = str(row[1]).strip() if row[1] else ''
        if not har or har == 'None': continue
        if not nom or nom == 'None': continue
        if nom.lower().startswith('номенклатура'): continue
        k = build_key(nom, har)
        m[k] = m.get(k, 0) + abs(to_num(row[3]))
    print(f'    → {len(m)} ключей')
    return m

def read_rezerv_obsh():
    print('  rezervy_obsh.xlsx ...')
    ws = openpyxl.load_workbook(FOLDER / 'rezervy_obsh.xlsx', data_only=True).active
    m = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        nom = str(row[0]).strip() if row[0] else ''
        if not nom or nom == 'None': continue
        k = build_key(nom, row[2])
        m[k] = m.get(k, 0) + to_num(row[10])
    print(f'    → {len(m)} ключей')
    return m

def process_sheet(ws, header_row, key_index,
                  vypusk, zakazano, otgr, ostatok,
                  shts, rub, rez_la, rez_obsh, old_itogo):

    def fc(col_name, data_map):
        ci = find_col(ws, header_row, col_name)
        if ci is None:
            print(f'    [!] не найдена: «{col_name}»')
            return
        n = fill_column(ws, ci, data_map, key_index)
        print(f'    «{col_name}»: {n} ячеек')

    fc('Поступление цех, шт',  vypusk)
    fc('Заказ ОПТ, шт',         zakazano)
    fc('Отгружено ОПТ, шт',     otgr)
    fc('Остатки 01.04 шт',       ostatok)

    itogo_ci = find_col(ws, header_row, 'Продажи ИТОГО, шт')
    if itogo_ci is None:
        print('    [!] «Продажи ИТОГО, шт» не найдена — пропускаем вставку')
    else:
        ws.insert_cols(itogo_ci)
        ws.cell(row=header_row, column=itogo_ci).value = NEW_SALES_COL
        n = fill_column(ws, itogo_ci, shts, key_index)
        print(f'    «{NEW_SALES_COL}»: {n} ячеек')

        rub_ci = find_col(ws, header_row, 'Продажи 02.04.-08.04. руб')
        if rub_ci:
            n = fill_column(ws, rub_ci, rub, key_index)
            print(f'    «Продажи 02.04.-08.04. руб»: {n} ячеек')
        else:
            print('    [!] «Продажи 02.04.-08.04. руб» не найдена')

        new_itogo_ci = find_col(ws, header_row, 'Продажи ИТОГО, шт')
        if new_itogo_ci:
            weekly_cols = []
            for ci in range(1, ws.max_column + 1):
                v = ws.cell(row=header_row, column=ci).value
                if v and WEEKLY_RE.match(str(v).strip()):
                    weekly_cols.append(ci)
            weekly_cols.sort()
            if weekly_cols:
                for ri in key_index.values():
                    refs = '+'.join(f'{get_column_letter(c)}{ri}' for c in weekly_cols)
                    ws.cell(row=ri, column=new_itogo_ci).value = f'={refs}'
                print(f'    «Продажи ИТОГО, шт»: формула ({len(weekly_cols)} нед.)')

    fc('Резерв LA 01.04 шт', rez_la)
    fc('Резерв  01.04 шт',   rez_obsh)

    abc_ci = find_col(ws, header_row, 'ABC-анализ')
    if abc_ci is None:
        print('    [!] «ABC-анализ» не найдена')
        return

    items = []
    for key, ri in key_index.items():
        total = max(0.0, old_itogo.get(key, 0.0) + shts.get(key, 0.0))
        items.append((ri, total))

    items.sort(key=lambda x: x[1], reverse=True)
    grand = sum(v for _, v in items)
    cum = 0.0
    row_cat = {}
    for ri, val in items:
        cum += val
        pct = cum / grand if grand > 0 else 1.0
        row_cat[ri] = 'A' if pct <= 0.80 else ('B' if pct <= 0.95 else 'C')

    for key, ri in key_index.items():
        cat = row_cat.get(ri, 'C')
        cell = ws.cell(row=ri, column=abc_ci)
        cell.value = cat
        cell.fill = FILL_A if cat == 'A' else (FILL_B if cat == 'B' else FILL_C)

    cnt = {c: sum(1 for v in row_cat.values() if v == c) for c in 'ABC'}
    print(f'    ABC: A={cnt["A"]}, B={cnt["B"]}, C={cnt["C"]}')


def main():
    print('=' * 55)
    print('EO Report Updater')
    print('=' * 55)

    required = [
        'EO_template.xlsx', 'vypusk_shi.xlsx', 'opt.xlsx',
        'ostatok.xlsx', 'prodazhi_nedelya.xlsx',
        'rezerv_lamoda.xlsx', 'rezervy_obsh.xlsx',
    ]
    missing = [f for f in required if not (FOLDER / f).exists()]
    if missing:
        print('❌ Файлы не найдены:')
        for f in missing: print(f'   • {f}')
        print(f'\nВсе файлы должны лежать рядом со скриптом:\n{FOLDER}')
        sys.exit(1)

    print('\n[1/3] Читаем источники...')
    vypusk         = read_vypusk()
    zakazano, otgr = read_opt()
    ostatok        = read_ostatok()
    shts, rub      = read_prodazhi()
    rez_la         = read_rezerv_lamoda()
    rez_obsh       = read_rezerv_obsh()

    print('\n  Читаем текущие «Продажи ИТОГО» для ABC...')
    old_itogo_all = {}
    wb_ro = openpyxl.load_workbook(TEMPLATE, data_only=True)
    for sn in SHEETS:
        if sn not in wb_ro.sheetnames: continue
        ws_ro = wb_ro[sn]
        try:
            hrow = get_header_row(ws_ro)
            kc   = find_col(ws_ro, hrow, 'Номенклатура+характеристика')
            ic   = find_col(ws_ro, hrow, 'Продажи ИТОГО, шт')
            d = {}
            if kc and ic:
                for ri in range(hrow + 1, ws_ro.max_row + 1):
                    k = ws_ro.cell(row=ri, column=kc).value
                    v = ws_ro.cell(row=ri, column=ic).value
                    if k: d[str(k).strip()] = to_num(v)
            old_itogo_all[sn] = d
            print(f'    {sn}: {len(d)} значений')
        except Exception as e:
            print(f'    [!] {sn}: {e}')
            old_itogo_all[sn] = {}
    wb_ro.close()

    print(f'\n[2/3] Обрабатываем шаблон...')
    wb = openpyxl.load_workbook(TEMPLATE, keep_vba=False)

    for sn in SHEETS:
        if sn not in wb.sheetnames:
            print(f'\n  [!] Лист не найден: {sn}')
            continue
        print(f'\n  ── {sn} ──')
        ws = wb[sn]
        try:
            hrow      = get_header_row(ws)
            kc        = find_col(ws, hrow, 'Номенклатура+характеристика')
            key_index = build_key_index(ws, hrow, kc)
            print(f'    позиций: {len(key_index)}, заголовки в строке {hrow}')
            process_sheet(ws, hrow, key_index,
                          vypusk, zakazano, otgr, ostatok,
                          shts, rub, rez_la, rez_obsh,
                          old_itogo_all.get(sn, {}))
        except Exception as e:
            print(f'  [!] Ошибка на листе {sn}: {e}')
            import traceback; traceback.print_exc()

    print(f'\n[3/3] Сохраняем {OUTPUT.name}...')
    wb.save(OUTPUT)
    print(f'\n✅ Готово!\n   {OUTPUT}')


if __name__ == '__main__':
    main()
